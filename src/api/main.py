"""FastAPI application for CamAI Attendance."""
from __future__ import annotations

import io
import platform
import re
import time
from collections import deque
from contextlib import asynccontextmanager
from datetime import date, datetime
from pathlib import Path
from typing import Literal

import cv2
import numpy as np
from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session, sessionmaker

try:
    from api.attendance_service import AttendanceService, Employee, get_engine, AttendanceLog
    from config import CAMERA_HEIGHT, CAMERA_SOURCE, CAMERA_WIDTH, DATA_RAW_DIR, get_runtime_settings, save_runtime_settings
    from detection.face_detector import FaceDetector
    from recognition.recognizer import FaceRecognizer
except ImportError:  # pragma: no cover
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from api.attendance_service import AttendanceService, Employee, get_engine, AttendanceLog
    from config import CAMERA_HEIGHT, CAMERA_SOURCE, CAMERA_WIDTH, DATA_RAW_DIR, get_runtime_settings, save_runtime_settings
    from detection.face_detector import FaceDetector
    from recognition.recognizer import FaceRecognizer


EMPLOYEE_ID_RE = re.compile(r"^[A-Za-z0-9_-]{2,32}$")
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
MAX_UPLOAD_BYTES = 8 * 1024 * 1024

engine = None
SessionLocal = None
detector: FaceDetector | None = None
recognizer: FaceRecognizer | None = None


class ManualAttendanceRequest(BaseModel):
    employee_id: str = Field(..., min_length=2, max_length=32)
    event_type: Literal["check_in", "check_out"]
    camera_id: str = "MANUAL"


class AttendanceRecordRequest(BaseModel):
    employee_id: str = Field(..., min_length=2, max_length=32)
    confidence: float
    camera_id: str = "REALTIME_CAMERA"
    is_real: bool = True
    liveness_score: float = 1.0
    event_type: Literal["check_in", "check_out"] | None = None
    emotion: str = "neutral"


class ConnectionManager:
    def __init__(self):
        self.active_connections: list[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in list(self.active_connections):
            try:
                await connection.send_json(message)
            except Exception:
                self.disconnect(connection)


manager = ConnectionManager()


@asynccontextmanager
async def lifespan(app: FastAPI):
    global engine, SessionLocal, detector, recognizer
    engine = get_engine()
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    detector = FaceDetector()
    recognizer = FaceRecognizer()
    print(
        "[CamAI] ready: "
        f"detector={detector.active_detector}, embeddings={detector.embedder is not None}, "
        f"face_db={recognizer.total}"
    )
    yield


app = FastAPI(
    title="CamAI Attendance API",
    description="Realtime face recognition attendance service",
    version="1.1.0",
    lifespan=lifespan,
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception:
        manager.disconnect(websocket)


def get_db():
    if SessionLocal is None:
        raise HTTPException(status_code=503, detail="Database is not initialized")
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _require_models() -> tuple[FaceDetector, FaceRecognizer]:
    if detector is None or recognizer is None:
        raise HTTPException(status_code=503, detail="AI models are not initialized")
    return detector, recognizer


def _validate_employee_id(employee_id: str) -> str:
    employee_id = employee_id.strip()
    if not EMPLOYEE_ID_RE.fullmatch(employee_id):
        raise HTTPException(
            status_code=400,
            detail="employee_id must be 2-32 chars: letters, numbers, underscore or dash",
        )
    return employee_id


async def _read_image_upload(file: UploadFile) -> np.ndarray:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only JPG, PNG, or WEBP images are accepted")

    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Image is too large")

    arr = np.frombuffer(data, np.uint8)
    frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if frame is None:
        raise HTTPException(status_code=400, detail="Cannot decode image")
    return frame


def _open_camera_capture():
    if isinstance(CAMERA_SOURCE, int):
        candidates = [CAMERA_SOURCE, 1, 2] if CAMERA_SOURCE == 0 else [CAMERA_SOURCE]
        backends = [cv2.CAP_ANY]
        if platform.system() == "Windows":
            backends = [cv2.CAP_DSHOW, cv2.CAP_MSMF, cv2.CAP_ANY]
    else:
        candidates = [CAMERA_SOURCE]
        backends = [cv2.CAP_FFMPEG, cv2.CAP_ANY]

    for source in candidates:
        for backend in backends:
            cap = cv2.VideoCapture(source, backend)
            if not cap.isOpened():
                cap.release()
                continue
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, CAMERA_WIDTH)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, CAMERA_HEIGHT)

            for _ in range(5):
                cap.read()
            ok, frame = cap.read()
            if ok and frame is not None:
                return cap
            cap.release()
    return None


def _draw_face_label(frame: np.ndarray, bbox: list[int], label: str, confidence: float, known: bool) -> None:
    color = (0, 220, 80) if known else (0, 80, 220)
    x1, y1, x2, y2 = bbox
    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 3)

    text = f"{label} {confidence:.0%}" if known else "unknown"
    scale = 0.7
    thickness = 2
    (tw, th), _ = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, scale, thickness)

    label_bg_padding = 6
    label_y = max(0, y1 - th - label_bg_padding * 2)
    cv2.rectangle(frame, (x1 - label_bg_padding, label_y), (x1 + tw + label_bg_padding, y1), color, -1)
    cv2.putText(
        frame,
        text,
        (x1, y1 - label_bg_padding),
        cv2.FONT_HERSHEY_SIMPLEX,
        scale,
        (255, 255, 255),
        thickness,
        cv2.LINE_AA,
    )




def _draw_detection_history(frame: np.ndarray, history: deque, panel_width: int, h: int) -> None:
    """Draw detection history on left panel (doesn't overlap camera)."""
    date_str = datetime.now().strftime("%Y/%m/%d")
    cv2.putText(
        frame,
        date_str,
        (12, 28),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.6,
        (255, 255, 255),
        1,
        cv2.LINE_AA,
    )

    cv2.line(frame, (8, 36), (panel_width - 8, 36), (80, 80, 100), 1)

    y_pos = 52
    row_height = 32

    for detection in list(history)[:8]:
        if y_pos + row_height > h - 15:
            break

        display_name = detection["display_name"]
        timestamp = detection["timestamp"].strftime("%H:%M:%S")

        cv2.putText(
            frame,
            display_name,
            (15, y_pos),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.55,
            (0, 220, 80),
            1,
            cv2.LINE_AA,
        )

        cv2.putText(
            frame,
            timestamp,
            (15, y_pos + 20),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.45,
            (150, 150, 180),
            1,
            cv2.LINE_AA,
        )

        cv2.line(frame, (10, y_pos + 24), (panel_width - 10, y_pos + 24), (50, 50, 70), 1)
        y_pos += row_height


def _camera_frame_generator(record_attendance: bool = True, every_n_frames: int = 2):
    det, rec_model = _require_models()
    if SessionLocal is None:
        raise HTTPException(status_code=503, detail="Database is not initialized")

    cap = _open_camera_capture()
    if cap is None:
        raise HTTPException(status_code=503, detail="Cannot open camera source")

    db = SessionLocal()
    svc = AttendanceService(db)
    frame_index = 0
    last_results: list[dict] = []
    detection_history: deque = deque(maxlen=10)
    displayed_identities: dict[str, datetime] = {}
    panel_width = 180
    latencies_ms: list[float] = []

    try:
        while True:
            started = time.perf_counter()
            ok, frame = cap.read()
            if not ok or frame is None:
                time.sleep(0.05)
                continue

            frame_index += 1
            if frame_index % max(1, every_n_frames) == 0:
                faces = det.detect(frame)
                last_results = []
                for face in faces:
                    rec = rec_model.recognize(face.embedding)
                    identity = rec["identity"]
                    known = identity != "unknown"
                    display_name = identity
                    attendance = None

                    if known:
                        employee = db.get(Employee, identity)
                        display_name = employee.name if employee else identity
                        if record_attendance:
                            attendance = svc.record(
                                employee_id=identity,
                                confidence=rec["confidence"],
                                camera_id="LIVE_CAMERA",
                                is_real=True,
                            )

                        if identity not in displayed_identities:
                            displayed_identities[identity] = datetime.now()
                            detection_history.append({
                                "identity": identity,
                                "display_name": display_name,
                                "timestamp": datetime.now(),
                            })

                    last_results.append(
                        {
                            "bbox": [int(value) for value in face.bbox.tolist()],
                            "identity": identity,
                            "display_name": display_name,
                            "confidence": rec["confidence"],
                            "known": known,
                            "attendance": attendance,
                        }
                    )

            h, w = frame.shape[:2]
            output = np.zeros((h, w + panel_width, 3), dtype=np.uint8)
            output[:, :panel_width] = [25, 25, 35]
            output[:, panel_width:] = frame

            for result in last_results:
                bbox = result["bbox"]
                adjusted_bbox = [
                    int(bbox[0]) + panel_width,
                    int(bbox[1]),
                    int(bbox[2]) + panel_width,
                    int(bbox[3]),
                ]
                _draw_face_label(
                    output,
                    adjusted_bbox,
                    result["display_name"],
                    result["confidence"],
                    result["known"],
                )

            latency_ms = (time.perf_counter() - started) * 1000
            latencies_ms.append(latency_ms)
            latencies_ms = latencies_ms[-30:]

            if last_results or frame_index % 30 == 0:
                avg_ms = sum(latencies_ms) / len(latencies_ms) if latencies_ms else 1
                cv2.putText(
                    output,
                    f"FPS: {1000 / max(0.001, avg_ms):.1f}",
                    (max(panel_width + 10, w + panel_width - 150), h - 12),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.5,
                    (200, 200, 200),
                    1,
                )

            _draw_detection_history(output, detection_history, panel_width, h)

            ok, encoded = cv2.imencode(".jpg", output, [int(cv2.IMWRITE_JPEG_QUALITY), 82])
            if not ok:
                continue
            yield (
                b"--frame\r\n"
                b"Content-Type: image/jpeg\r\n\r\n"
                + encoded.tobytes()
                + b"\r\n"
            )
    finally:
        db.close()
        cap.release()


@app.get("/api/stream", tags=["System"])
def stream_camera():
    return StreamingResponse(
        _camera_frame_generator(record_attendance=True),
        media_type="multipart/x-mixed-replace; boundary=frame"
    )


@app.get("/health", tags=["System"])
def health():
    return {
        "status": "ok",
        "detector": detector.active_detector if detector else None,
        "embeddings_enabled": bool(detector and detector.embedder is not None),
        "db_size": recognizer.total if recognizer else 0,
        "people": recognizer.people if recognizer else [],
    }


@app.post("/api/recognize", tags=["Recognition"])
async def recognize(
    file: UploadFile = File(..., description="Camera frame as JPG/PNG/WEBP"),
    camera_id: str = "CAM_01",
    db: Session = Depends(get_db),
):
    det, rec_model = _require_models()
    frame = await _read_image_upload(file)
    faces = det.detect(frame)
    if not faces:
        return {"status": "no_face", "face_count": 0, "results": []}

    svc = AttendanceService(db)
    results = []
    for face in faces:
        rec = rec_model.recognize(face.embedding)
        attendance = None
        if rec["identity"] != "unknown":
            attendance = svc.record(
                employee_id=rec["identity"],
                confidence=rec["confidence"],
                camera_id=camera_id,
                is_real=True,
            )
        results.append({**rec, "attendance": attendance, "bbox": face.bbox.tolist()})

    return {"status": "ok", "face_count": len(faces), "results": results}


@app.post("/api/employees/register", tags=["Employees"])
async def register(
    employee_id: str = Form(...),
    name: str = Form(...),
    department: str = Form(""),
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    det, rec_model = _require_models()
    employee_id = _validate_employee_id(employee_id)
    if len(files) < 3:
        raise HTTPException(status_code=400, detail="Upload at least 3 clear face images")

    save_dir = Path(DATA_RAW_DIR) / employee_id
    save_dir.mkdir(parents=True, exist_ok=True)

    new_embeddings = []
    saved = 0
    for idx, upload in enumerate(files):
        frame = await _read_image_upload(upload)
        detected = det.detect(frame)
        if not detected or detected[0].embedding is None:
            continue

        suffix = Path(upload.filename or ".jpg").suffix.lower()
        path = save_dir / f"{employee_id}_{idx:03d}{suffix}"
        _, encoded = cv2.imencode(suffix if suffix != ".jpg" else ".jpg", frame)
        with open(path, "wb") as file:
            file.write(encoded.tobytes())

        new_embeddings.append(detected[0].embedding)
        saved += 1

    if not new_embeddings:
        raise HTTPException(status_code=400, detail="No usable face embeddings found in upload")

    added = rec_model.add_person(employee_id, new_embeddings)
    emp = Employee(id=employee_id, name=name.strip(), department=department.strip(), active=True)
    db.merge(emp)
    db.commit()

    return {
        "status": "success",
        "employee_id": employee_id,
        "images_saved": saved,
        "embeddings_added": added,
        "total_in_db": rec_model.total,
    }


@app.delete("/api/employees/{employee_id}", tags=["Employees"])
def remove_employee(employee_id: str, db: Session = Depends(get_db)):
    _, rec_model = _require_models()
    employee_id = _validate_employee_id(employee_id)
    removed = rec_model.remove_person(employee_id)
    emp = db.get(Employee, employee_id)
    if emp:
        emp.active = False
        db.commit()

    # Clean up raw images folder to free space
    import shutil
    save_dir = Path(DATA_RAW_DIR) / employee_id
    if save_dir.exists():
        try:
            shutil.rmtree(save_dir)
        except Exception:
            pass

    return {"status": "removed", "employee_id": employee_id, "embeddings_removed": removed}


@app.get("/api/employees", tags=["Employees"])
def list_employees(db: Session = Depends(get_db)):
    employees = db.query(Employee).filter(Employee.active == True).order_by(Employee.id).all()  # noqa: E712
    return [
        {"id": employee.id, "name": employee.name, "department": employee.department}
        for employee in employees
    ]


@app.get("/api/attendance/report", tags=["Attendance"])
def attendance_report(
    target_date: str | None = None,
    format: str = "json",
    db: Session = Depends(get_db),
):
    try:
        report_date = date.fromisoformat(target_date) if target_date else date.today()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="target_date must use YYYY-MM-DD") from exc

    report = AttendanceService(db).get_daily_report(report_date)
    if format != "excel":
        return JSONResponse({"date": str(report_date), "data": report})

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Attendance_{report_date}"
    headers = [
        "Employee ID",
        "Name",
        "Department",
        "Check In",
        "Check Out",
        "Hours",
        "Late",
        "Status",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = Alignment(horizontal="center")

    for row in report:
        sheet.append(
            [
                row["employee_id"],
                row["name"],
                row.get("department", ""),
                row["check_in"] or "-",
                row["check_out"] or "-",
                row["work_hours"] or "-",
                "Yes" if row["late"] else "No",
                row["status"],
            ]
        )

    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 4
        sheet.column_dimensions[column[0].column_letter].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_{report_date}.xlsx"},
    )


@app.post("/api/attendance/record", tags=["Attendance"])
async def record_attendance_api(payload: AttendanceRecordRequest, db: Session = Depends(get_db)):
    employee_id = _validate_employee_id(payload.employee_id)
    if payload.event_type is not None:
        result = AttendanceService(db).record_event(
            employee_id=employee_id,
            event_type=payload.event_type,
            confidence=payload.confidence,
            camera_id=payload.camera_id,
            is_real=payload.is_real,
            liveness_score=payload.liveness_score,
            emotion=payload.emotion,
        )
    else:
        result = AttendanceService(db).record(
            employee_id=employee_id,
            confidence=payload.confidence,
            camera_id=payload.camera_id,
            is_real=payload.is_real,
            liveness_score=payload.liveness_score,
            emotion=payload.emotion,
        )
    if result.get("logged"):
        await manager.broadcast({
            "type": "attendance_event",
            "data": result
        })
    return result


@app.post("/api/attendance/manual", tags=["Attendance"])
async def manual_attendance(payload: ManualAttendanceRequest, db: Session = Depends(get_db)):
    employee_id = _validate_employee_id(payload.employee_id)
    result = AttendanceService(db).record_event(
        employee_id=employee_id,
        event_type=payload.event_type,
        confidence=1.0,
        camera_id=payload.camera_id,
        is_real=True,
        liveness_score=1.0,
    )
    if not result.get("logged"):
        return JSONResponse(status_code=409, content=result)
    
    await manager.broadcast({
        "type": "attendance_event",
        "data": result
    })
    return result


@app.get("/api/attendance/logs", tags=["Attendance"])
def list_recent_logs(employee_id: str | None = None, limit: int = 50, db: Session = Depends(get_db)):
    limit = max(1, min(limit, 100))
    query = db.query(AttendanceLog)
    if employee_id:
        query = query.filter(AttendanceLog.employee_id == employee_id)

    logs = query.order_by(AttendanceLog.timestamp.desc()).limit(limit).all()

    result = []
    for log in logs:
        employee = db.get(Employee, log.employee_id)
        result.append({
            "employee_id": log.employee_id,
            "name": employee.name if employee else log.employee_id,
            "department": employee.department if employee else "",
            "event_type": log.event_type,
            "timestamp": log.timestamp.isoformat(),
            "time": log.timestamp.strftime("%H:%M:%S"),
            "date": log.timestamp.strftime("%Y-%m-%d"),
            "confidence": round(float(log.confidence or 0.0), 4),
            "camera_id": log.camera_id,
            "is_real_face": bool(log.is_real_face if log.is_real_face is not None else True),
            "liveness_score": round(float(log.liveness_score if log.liveness_score is not None else 1.0), 4),
            "emotion": str(log.emotion) if log.emotion else "neutral",
        })
    return result


class SettingsRequest(BaseModel):
    late_threshold: str
    cooldown_minutes: int
    face_threshold: float


@app.get("/api/settings", tags=["System"])
def get_settings():
    return get_runtime_settings()


@app.post("/api/settings", tags=["System"])
def update_settings(payload: SettingsRequest):
    if not re.match(r"^\d{2}:\d{2}$", payload.late_threshold):
        raise HTTPException(status_code=400, detail="late_threshold must be in HH:MM format")
    if not (1 <= payload.cooldown_minutes <= 120):
        raise HTTPException(status_code=400, detail="cooldown_minutes must be between 1 and 120")
    if not (0.1 <= payload.face_threshold <= 1.0):
        raise HTTPException(status_code=400, detail="face_threshold must be between 0.1 and 1.0")

    save_runtime_settings({
        "late_threshold": payload.late_threshold,
        "cooldown_minutes": payload.cooldown_minutes,
        "face_threshold": payload.face_threshold
    })
    return {"status": "success", "settings": get_runtime_settings()}


# Serve the frontend statically at the root path for easy Docker/Local deployment
try:
    from fastapi.staticfiles import StaticFiles
    frontend_path = Path(__file__).resolve().parents[2] / "frontend"
    if frontend_path.exists():
        app.mount("/", StaticFiles(directory=str(frontend_path), html=True), name="frontend")
except Exception as e:
    print(f"[Warning] Failed to mount static frontend: {e}")

